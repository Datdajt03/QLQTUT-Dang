# -*- coding: utf-8 -*-
# Flask API cho xuat Excel & PDF – He thong Ket nap Dang
# pip install -r requirements.txt
import sys, io
# Force UTF-8 output on Windows (fix cp1252 UnicodeEncodeError)
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
if sys.stderr.encoding != 'utf-8':
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pymysql
import pymysql.cursors
import openpyxl
from openpyxl.styles import Font as XLFont, PatternFill as XLPatternFill, Alignment as XLAlignment, Border as XLBorder, Side as XLSide
from openpyxl.utils import get_column_letter

# ReportLab imports for PDF
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import io
import os
from datetime import datetime, date

app = Flask(__name__)
CORS(app)

# ─── Database config ──────────────────────────────────────────────────────────
DB_CFG = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'db': 'quan_ly_ket_nap_dang',
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor,
    'autocommit': True,
}

def get_db():
    return pymysql.connect(**DB_CFG)

# ─── Constants & Styles ───────────────────────────────────────────────────────
RED   = 'C8102E'
GOLD  = 'FFD700'
WHITE = 'FFFFFF'
LIGHT = 'FFF3F3'
STRIPE= 'FFF8F8'

HEADERS = [
    'STT','Mã GV/SV','Họ và tên','SĐT','Giới tính','Ngày sinh','Dân tộc',
    'Quê quán','Chức vụ','Lớp','Chi bộ công nhận','Số BC cảm tình Đảng',
    'Ngày họp CB công nhận','Đảng viên giúp đỡ','Ngày phân công giúp đỡ',
    'Số QĐ mở lớp BD','Ngày QĐ mở lớp','Thời gian lớp BD','Ngày cấp CC',
    'Số QĐ CC BD','Đơn vị cấp CC','ĐV công tác khi cấp CC',
    'CB sinh hoạt khi cấp CC','Đảng uỷ khi cấp CC','Tỉnh uỷ khi cấp CC',
    'Mã số','Kết nạp Đảng','Ngày quyết định','Số QĐ kết nạp','Ngày kết nạp',
    'ĐV hướng dẫn','Ngày chuyển SH','Nơi chuyển tới','Trạng thái',
]

FIELDS = [
    'ma_gvsv','ho_ten','sdt','gioi_tinh','ngay_sinh','dan_toc','que_quan',
    'chuc_vu','lop','chi_bo_cong_nhan','so_bc_cam_tinh','ngay_hop_cam_tinh',
    'dang_vien_giup_do','ngay_phan_cong_giup_do','so_qd_mo_lop',
    'ngay_qd_mo_lop','tg_lop_boi_duong','ngay_cap_cc','so_qd_cc',
    'don_vi_cap_cc','ten_dv_congtac_khi_cap_cc','ten_chibo_khi_cap_cc',
    'ten_danguy_khi_cap_cc','ten_tinhuy_khi_cap_cc','ma_so','ket_nap_dang',
    'ngay_quyet_dinh','so_qd_ket_nap','ngay_ket_nap','dang_vien_huong_dan',
    'ngay_chuyen_sinh_hoat','noi_chuyen_toi','trang_thai',
]

DATE_FIELDS = {
    'ngay_sinh','ngay_hop_cam_tinh','ngay_phan_cong_giup_do','ngay_qd_mo_lop',
    'ngay_cap_cc','ngay_quyet_dinh','ngay_ket_nap','ngay_chuyen_sinh_hoat',
}

COL_WIDTHS = [
    6,12,28,14,10,14,14,28,16,22,28,22,18,22,18,
    18,16,28,16,16,22,28,22,22,22,12,18,16,18,16,
    22,16,25,16,
]

# Register Fonts for PDF (Use Windows Times New Roman for Vietnamese support)
FONT_REGULAR = 'Times'
FONT_BOLD = 'Times-Bold'
FONT_ITALIC = 'Times-Italic'

win_fonts_dir = "C:\\Windows\\Fonts"
if os.path.exists(win_fonts_dir):
    try:
        pdfmetrics.registerFont(TTFont('TimesNewRoman', os.path.join(win_fonts_dir, 'times.ttf')))
        pdfmetrics.registerFont(TTFont('TimesNewRoman-Bold', os.path.join(win_fonts_dir, 'timesbd.ttf')))
        pdfmetrics.registerFont(TTFont('TimesNewRoman-Italic', os.path.join(win_fonts_dir, 'timesi.ttf')))
        pdfmetrics.registerFont(TTFont('TimesNewRoman-BoldItalic', os.path.join(win_fonts_dir, 'timesbi.ttf')))
        FONT_REGULAR = 'TimesNewRoman'
        FONT_BOLD = 'TimesNewRoman-Bold'
        FONT_ITALIC = 'TimesNewRoman-Italic'
        print("[OK] Font TimesNewRoman da dang ky thanh cong!")
    except Exception as e:
        print(f"[WARN] Loi dang ky font: {e}. Su dung font tieu chuan thay the.")

# ─── Helpers ──────────────────────────────────────────────────────────────────
def fmt_date(v):
    if not v: return ''
    if isinstance(v, (date,)): return v.strftime('%d/%m/%Y')
    return str(v)

def serialize_row(r):
    out = {}
    for k, v in r.items():
        if isinstance(v, (date,)): out[k] = fmt_date(v)
        elif v is None: out[k] = ''
        else: out[k] = v
    return out

def thin_border():
    s = XLSide(style='thin', color='CCCCCC')
    return XLBorder(left=s, right=s, top=s, bottom=s)

def make_header_style():
    return {
        'font': XLFont(name='Times New Roman', bold=True, size=10, color=WHITE),
        'fill': XLPatternFill(start_color=RED, end_color=RED, fill_type='solid'),
        'align': XLAlignment(horizontal='center', vertical='center', wrap_text=True),
        'border': thin_border(),
    }

# ─── Excel builder (Type 1) ──────────────────────────────────────────────────
def build_list_workbook(rows: list, subtitle: str = '') -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Danh sách'

    n_cols = len(HEADERS)
    last_col = get_column_letter(n_cols)

    # Title
    ws.merge_cells(f'A1:{last_col}1')
    c = ws['A1']
    c.value = 'DANH SÁCH QUẦN CHÚNG ƯU TÚ PHỤC VỤ KẾT NẠP ĐẢNG'
    c.font = XLFont(name='Times New Roman', bold=True, size=14, color=RED)
    c.alignment = XLAlignment(horizontal='center', vertical='center')
    c.fill = XLPatternFill(start_color=LIGHT, end_color=LIGHT, fill_type='solid')
    ws.row_dimensions[1].height = 32

    # Subtitle
    ws.merge_cells(f'A2:{last_col}2')
    sub = subtitle or ''
    ws['A2'].value = (f'{sub}  |  ' if sub else '') + f'Xuất ngày: {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  Tổng: {len(rows)} người'
    ws['A2'].font = XLFont(name='Times New Roman', italic=True, size=10, color='555555')
    ws['A2'].alignment = XLAlignment(horizontal='center')
    ws.row_dimensions[2].height = 18

    # Divider line
    ws.merge_cells(f'A3:{last_col}3')
    ws['A3'].fill = XLPatternFill(start_color=RED, fill_type='solid')
    ws.row_dimensions[3].height = 4

    # Header
    hs = make_header_style()
    ws.append(HEADERS)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=4, column=col)
        cell.font  = hs['font']
        cell.fill  = hs['fill']
        cell.alignment = hs['align']
        cell.border = hs['border']
    ws.row_dimensions[4].height = 40

    # Rows
    border = thin_border()
    for i, row in enumerate(rows, 1):
        data = [i]
        for f in FIELDS[1:]:
            v = row.get(f) or ''
            if f in DATE_FIELDS: v = fmt_date(v)
            data.append(str(v) if v else '')
        data.insert(1, str(row.get('ma_gvsv') or ''))
        ws.append(data)

        dr = 4 + i
        fill_color = WHITE if i % 2 == 1 else STRIPE
        fill = XLPatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=dr, column=col)
            cell.font   = XLFont(name='Times New Roman', size=10)
            cell.fill   = fill
            cell.border = border
            cell.alignment = XLAlignment(vertical='center', wrap_text=True)
        ws.cell(row=dr, column=1).alignment = XLAlignment(horizontal='center', vertical='center')
        ws.row_dimensions[dr].height = 20

    # Summary row
    summary_row = 4 + len(rows) + 1
    ws.merge_cells(f'A{summary_row}:{last_col}{summary_row}')
    sc = ws[f'A{summary_row}']
    sc.value = f'Tổng số: {len(rows)} người'
    sc.font = XLFont(name='Times New Roman', bold=True, size=11, color=RED)
    sc.alignment = XLAlignment(horizontal='right', vertical='center', indent=2)
    sc.fill = XLPatternFill(start_color=LIGHT, fill_type='solid')
    ws.row_dimensions[summary_row].height = 24

    for i, w in enumerate(COL_WIDTHS[:n_cols], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A5'
    return wb

# ─── PDF Builders (Types 2 & 3) ──────────────────────────────────────────────
def build_profile_pdf(row: dict) -> io.BytesIO:
    """Tạo file PDF chi tiết hồ sơ lý lịch cá nhân (Loại 2)"""
    buf = io.BytesIO()
    # Margins: 0.5 inch (36 points)
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    story = []

    # Styles
    title_style = ParagraphStyle('Title', fontName=FONT_BOLD, fontSize=16, leading=20, alignment=1, textColor=colors.HexColor('#C8102E'))
    subtitle_style = ParagraphStyle('Sub', fontName=FONT_ITALIC, fontSize=10, leading=12, alignment=1, textColor=colors.HexColor('#555555'))
    section_style = ParagraphStyle('Section', fontName=FONT_BOLD, fontSize=11, leading=14, textColor=colors.white)
    label_style = ParagraphStyle('Label', fontName=FONT_BOLD, fontSize=9, leading=12, textColor=colors.HexColor('#333333'))
    val_style = ParagraphStyle('Value', fontName=FONT_REGULAR, fontSize=9, leading=12, textColor=colors.black)

    # 1. Header & Title
    story.append(Paragraph("HỒ SƠ QUẦN CHÚNG ƯU TÚ PHỤC VỤ KẾT NẠP ĐẢNG", title_style))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"Tên đối tượng: <b>{row.get('ho_ten','')}</b>  |  Mã số: {row.get('ma_gvsv','—')}  |  Lớp: {row.get('lop','—')}", subtitle_style))
    story.append(Spacer(1, 10))

    # Divider
    div_table = Table([[""]], colWidths=[540])
    div_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#C8102E')),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ('TOPPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(div_table)
    story.append(Spacer(1, 12))

    # Avatar Block
    avatar_path = row.get('avatar')
    avatar_elem = None
    if avatar_path:
        full_avatar_path = os.path.join(dirname_dir(), avatar_path)
        if os.path.exists(full_avatar_path):
            try:
                # 80x100 point avatar box
                avatar_elem = RLImage(full_avatar_path, width=70, height=90)
            except:
                pass

    def make_section_header(title_text):
        tbl = Table([[Paragraph(title_text, section_style)]], colWidths=[540])
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#C8102E')),
            ('PADDING', (0,0), (-1,-1), 5),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ]))
        return tbl

    def make_data_table(fields_list, bg_alt=True):
        data = []
        for i, (lbl, val) in enumerate(fields_list):
            val_p = Paragraph(str(val) if val else '—', val_style)
            lbl_p = Paragraph(lbl, label_style)
            data.append([lbl_p, val_p])

        tbl = Table(data, colWidths=[200, 340])
        t_style = [
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E0E0E0')),
            ('PADDING', (0,0), (-1,-1), 4),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]
        if bg_alt:
            for idx in range(len(fields_list)):
                bg = colors.HexColor('#FFFDFD') if idx % 2 == 0 else colors.HexColor('#FFF6F6')
                t_style.append(('BACKGROUND', (0, idx), (-1, idx), bg))
        tbl.setStyle(TableStyle(t_style))
        return tbl

    # Section 1
    story.append(make_section_header("I. THÔNG TIN CÁ NHÂN"))
    personal = [
        ("Mã GV/SV", row.get('ma_gvsv')),
        ("Giới tính", row.get('gioi_tinh')),
        ("Ngày sinh", fmt_date(row.get('ngay_sinh'))),
        ("Dân tộc", row.get('dan_toc')),
        ("Quê quán", row.get('que_quan')),
        ("Số điện thoại", row.get('sdt')),
        ("Chức vụ đoàn thể/lớp", row.get('chuc_vu')),
        ("Lớp sinh hoạt", row.get('lop')),
    ]
    
    if avatar_elem:
        # If avatar exists, display it side by side with the info table
        info_tbl = Table([
            [Paragraph("Mã GV/SV", label_style), Paragraph(str(row.get('ma_gvsv','—')), val_style)],
            [Paragraph("Giới tính", label_style), Paragraph(str(row.get('gioi_tinh','—')), val_style)],
            [Paragraph("Ngày sinh", label_style), Paragraph(fmt_date(row.get('ngay_sinh')), val_style)],
            [Paragraph("Dân tộc", label_style), Paragraph(str(row.get('dan_toc','—')), val_style)],
            [Paragraph("SĐT", label_style), Paragraph(str(row.get('sdt','—')), val_style)],
            [Paragraph("Chức vụ", label_style), Paragraph(str(row.get('chuc_vu','—')), val_style)],
            [Paragraph("Lớp", label_style), Paragraph(str(row.get('lop','—')), val_style)],
        ], colWidths=[110, 330])
        info_tbl.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E0E0E0')),
            ('PADDING', (0,0), (-1,-1), 3),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        
        # Merge avatar and table
        side_table = Table([[info_tbl, avatar_elem]], colWidths=[450, 90])
        side_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (1,0), (1,0), 'CENTER'),
            ('LEFTPADDING', (1,0), (1,0), 10),
            ('PADDING', (0,0), (-1,-1), 0),
        ]))
        story.append(side_table)
        
        # Add rest of Section 1
        rem_tbl = make_data_table([
            ("Quê quán", row.get('que_quan'))
        ], bg_alt=False)
        story.append(rem_tbl)
    else:
        story.append(make_data_table(personal))
    
    story.append(Spacer(1, 10))

    # Section 2
    story.append(make_section_header("II. CHI BỘ & CẢM TÌNH ĐẢNG"))
    sec2 = [
        ("Chi bộ công nhận cảm tình", row.get('chi_bo_cong_nhan')),
        ("Số Báo cáo Chi bộ công nhận CT", row.get('so_bc_cam_tinh')),
        ("Ngày họp Chi bộ thông qua", fmt_date(row.get('ngay_hop_cam_tinh'))),
        ("Đảng viên được phân công giúp đỡ", row.get('dang_vien_giup_do')),
        ("Ngày chi bộ ra quyết định phân công", fmt_date(row.get('ngay_phan_cong_giup_do'))),
    ]
    story.append(make_data_table(sec2))
    story.append(Spacer(1, 10))

    # Section 3
    story.append(make_section_header("III. LỚP BỒI DƯỠNG NHẬN THỨC VỀ ĐẢNG"))
    sec3 = [
        ("Số Quyết định mở lớp bồi dưỡng", row.get('so_qd_mo_lop')),
        ("Ngày ra Quyết định mở lớp", fmt_date(row.get('ngay_qd_mo_lop'))),
        ("Thời gian học lớp nhận thức Đảng", row.get('tg_lop_boi_duong')),
        ("Ngày cấp chứng chỉ bồi dưỡng", fmt_date(row.get('ngay_cap_cc'))),
        ("Số chứng chỉ bồi dưỡng", row.get('so_qd_cc')),
        ("Đơn vị/Trường cấp chứng chỉ", row.get('don_vi_cap_cc')),
        ("Tên đơn vị công tác khi cấp chứng chỉ", row.get('ten_dv_congtac_khi_cap_cc')),
        ("Chi bộ sinh hoạt khi cấp chứng chỉ", row.get('ten_chibo_khi_cap_cc')),
        ("Đảng ủy trực thuộc khi cấp chứng chỉ", row.get('ten_danguy_khi_cap_cc')),
        ("Tỉnh/Thành ủy sinh hoạt khi cấp CC", row.get('ten_tinhuy_khi_cap_cc')),
    ]
    story.append(make_data_table(sec3))
    story.append(Spacer(1, 10))

    # Section 4 & 5 Keep Together
    sec4_5 = []
    sec4_5.append(make_section_header("IV. QUYẾT ĐỊNH & THỜI GIAN KẾT NẠP"))
    sec4 = [
        ("Mã số hồ sơ Đảng viên", row.get('ma_so')),
        ("Chi tiết kết nạp Đảng", row.get('ket_nap_dang')),
        ("Ngày ra Quyết định kết nạp", fmt_date(row.get('ngay_quyet_dinh'))),
        ("Số Quyết định kết nạp Đảng viên", row.get('so_qd_ket_nap')),
        ("Ngày tổ chức lễ kết nạp", fmt_date(row.get('ngay_ket_nap'))),
        ("Đảng viên được giao hướng dẫn", row.get('dang_vien_huong_dan')),
    ]
    sec4_5.append(make_data_table(sec4))
    sec4_5.append(Spacer(1, 10))

    sec4_5.append(make_section_header("V. CHUYỂN SINH HOẠT VÀ TRẠNG THÁI"))
    sec5 = [
        ("Ngày chuyển sinh hoạt chính thức", fmt_date(row.get('ngay_chuyen_sinh_hoat'))),
        ("Nơi chuyển sinh hoạt đảng tới", row.get('noi_chuyen_toi')),
        ("Trạng thái theo dõi hiện tại", row.get('trang_thai')),
    ]
    sec4_5.append(make_data_table(sec5))
    
    if row.get('ghi_chu'):
        sec4_5.append(Spacer(1, 8))
        sec4_5.append(Table([[Paragraph("<b>Ghi chú:</b> " + row.get('ghi_chu'), val_style)]], colWidths=[540], 
                            style=TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
                                              ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#FCFCFC')),
                                              ('PADDING', (0,0), (-1,-1), 6)])))

    story.append(KeepTogether(sec4_5))

    # Footer/Date print
    story.append(Spacer(1, 15))
    foot_style = ParagraphStyle('Foot', fontName=FONT_ITALIC, fontSize=8, leading=10, alignment=2, textColor=colors.HexColor('#888888'))
    story.append(Paragraph(f"Tài liệu xuất ngày: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}  |  Phần mềm Quản lý kết nạp Đảng", foot_style))

    doc.build(story)
    buf.seek(0)
    return buf


def build_list_pdf(rows: list, subtitle: str = '') -> io.BytesIO:
    """Tạo file PDF dạng danh sách bảng ngang (Loại 3)"""
    buf = io.BytesIO()
    # Landscape A4 size: 842 x 595
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), leftMargin=28, rightMargin=28, topMargin=28, bottomMargin=28)
    story = []

    # Title Styles
    title_style = ParagraphStyle('T', fontName=FONT_BOLD, fontSize=15, leading=18, alignment=1, textColor=colors.HexColor('#C8102E'))
    sub_style = ParagraphStyle('S', fontName=FONT_ITALIC, fontSize=9, leading=11, alignment=1, textColor=colors.HexColor('#555555'))
    
    story.append(Paragraph("DANH SÁCH QUẦN CHÚNG ƯU TÚ PHỤC VỤ KẾT NẠP ĐẢNG", title_style))
    story.append(Spacer(1, 4))
    story.append(Paragraph((f"{subtitle}  |  " if subtitle else "") + f"Ngày in: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Tổng số: {len(rows)} quần chúng", sub_style))
    story.append(Spacer(1, 10))

    # Header & table styles
    th_style = ParagraphStyle('TH', fontName=FONT_BOLD, fontSize=8, leading=10, textColor=colors.white, alignment=1)
    td_style = ParagraphStyle('TD', fontName=FONT_REGULAR, fontSize=8, leading=10)
    td_bold  = ParagraphStyle('TDB', fontName=FONT_BOLD, fontSize=8, leading=10)
    
    # Selected columns for list table
    cols = ['STT', 'Mã GV/SV', 'Họ và tên', 'Giới tinh', 'Lớp học/Sinh hoạt', 'Chi bộ đề nghị', 'Đảng viên giúp đỡ', 'Ngày họp CB', 'Trạng thái']
    pdf_headers = [Paragraph(c, th_style) for c in cols]
    
    data = [pdf_headers]
    for idx, row in enumerate(rows, 1):
        data.append([
            Paragraph(str(idx), td_style),
            Paragraph(str(row.get('ma_gvsv') or '—'), td_style),
            Paragraph(f"<b>{row.get('ho_ten','')}</b>", td_style),
            Paragraph(str(row.get('gioi_tinh') or '—'), td_style),
            Paragraph(str(row.get('lop') or '—'), td_style),
            Paragraph(str(row.get('chi_bo_cong_nhan') or '—'), td_style),
            Paragraph(str(row.get('dang_vien_giup_do') or '—'), td_style),
            Paragraph(fmt_date(row.get('ngay_hop_cam_tinh')), td_style),
            Paragraph(str(row.get('trang_thai') or '—'), td_bold if row.get('trang_thai') == 'Đã kết nạp' else td_style)
        ])

    # Table widths sum = 786 pt (available landscape space is 792 pt)
    w = [30, 60, 140, 45, 120, 130, 110, 65, 86]
    t = Table(data, colWidths=w, repeatRows=1)
    
    t_style = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#C8102E')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
        ('PADDING', (0,0), (-1,-1), 4),
    ]
    for r_idx in range(1, len(data)):
        bg = colors.HexColor('#FFFFFF') if r_idx % 2 == 1 else colors.HexColor('#FFF9F9')
        t_style.append(('BACKGROUND', (0, r_idx), (-1, r_idx), bg))
    t.setStyle(TableStyle(t_style))
    
    story.append(t)
    story.append(Spacer(1, 10))
    story.append(Paragraph(f"Tổng cộng danh sách: {len(rows)} bản ghi.", ParagraphStyle('F', fontName=FONT_BOLD, fontSize=9, textColor=colors.HexColor('#C8102E'))))
    
    doc.build(story)
    buf.seek(0)
    return buf

# Helper to find dirname
def dirname_dir():
    return dirname_dir_path

dirname_dir_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ─── Routes ───────────────────────────────────────────────────────────────────
@app.route('/health')
def health():
    try:
        db = get_db()
        with db.cursor() as cur:
            cur.execute("SELECT COUNT(*) as n FROM doi_tuong")
            n = cur.fetchone()['n']
        db.close()
        return jsonify({'status': 'ok', 'total': n, 'version': '1.1'})
    except Exception as e:
        return jsonify({'status': 'error', 'msg': str(e)}), 500


@app.route('/api/filters')
def get_filters():
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT DISTINCT lop FROM doi_tuong WHERE lop IS NOT NULL AND lop!='' ORDER BY lop")
            lops = [r['lop'] for r in cur.fetchall()]
            cur.execute("SELECT DISTINCT chi_bo_cong_nhan FROM doi_tuong WHERE chi_bo_cong_nhan IS NOT NULL AND chi_bo_cong_nhan!='' ORDER BY chi_bo_cong_nhan")
            chibos = [r['chi_bo_cong_nhan'] for r in cur.fetchall()]
            cur.execute("SELECT COUNT(*) as n FROM doi_tuong")
            total = cur.fetchone()['n']
        return jsonify({'lops': lops, 'chibos': chibos, 'total': total})
    finally:
        db.close()


@app.route('/api/list')
def get_list():
    ft = request.args.get('filter_type', 'all')
    fv = request.args.get('filter_value', '')
    db = get_db()
    try:
        with db.cursor() as cur:
            if ft == 'lop' and fv:
                cur.execute("SELECT id,ma_gvsv,ho_ten,gioi_tinh,lop,chi_bo_cong_nhan,trang_thai,ngay_ket_nap,avatar FROM doi_tuong WHERE lop=%s ORDER BY ho_ten", (fv,))
            elif ft == 'chibo' and fv:
                cur.execute("SELECT id,ma_gvsv,ho_ten,gioi_tinh,lop,chi_bo_cong_nhan,trang_thai,ngay_ket_nap,avatar FROM doi_tuong WHERE chi_bo_cong_nhan=%s ORDER BY ho_ten", (fv,))
            else:
                cur.execute("SELECT id,ma_gvsv,ho_ten,gioi_tinh,lop,chi_bo_cong_nhan,trang_thai,ngay_ket_nap,avatar FROM doi_tuong ORDER BY ho_ten")
            rows = [serialize_row(r) for r in cur.fetchall()]
        return jsonify(rows)
    finally:
        db.close()


@app.route('/api/export/template')
def export_template():
    """Xuất file Excel mẫu chuẩn hóa có tiêu đề cột và ID tên cột tương ứng để nhập dữ liệu"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mau_Nhap_Du_Lieu'

    # Rows: Row 1 = DB Column ID (Hidden/Reference), Row 2 = Display Header Title
    db_col_ids = [
        'stt_dummy', 'ma_gvsv', 'ho_ten', 'sdt', 'gioi_tinh', 'ngay_sinh', 'dan_toc',
        'que_quan', 'chuc_vu', 'lop', 'chi_bo_cong_nhan', 'so_bc_cam_tinh',
        'ngay_hop_cam_tinh', 'dang_vien_giup_do', 'ngay_phan_cong_giup_do',
        'so_qd_mo_lop', 'ngay_qd_mo_lop', 'tg_lop_boi_duong', 'ngay_cap_cc',
        'so_qd_cc', 'don_vi_cap_cc', 'ten_dv_congtac_khi_cap_cc',
        'ten_chibo_khi_cap_cc', 'ten_danguy_khi_cap_cc', 'ten_tinhuy_khi_cap_cc',
        'ma_so', 'ket_nap_dang', 'ngay_quyet_dinh', 'so_qd_ket_nap', 'ngay_ket_nap',
        'dang_vien_huong_dan', 'ngay_chuyen_sinh_hoat', 'noi_chuyen_toi', 'trang_thai'
    ]

    # Row 1: DB Column IDs
    ws.append([f'[ID: {cid}]' for cid in db_col_ids])
    for col_idx in range(1, len(db_col_ids) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = XLFont(name='Times New Roman', bold=True, size=9, color='777777')
        cell.fill = XLPatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
        cell.alignment = XLAlignment(horizontal='center', vertical='center')

    # Row 2: Human Display Headers
    ws.append(HEADERS)
    header_style = make_header_style()
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = header_style['font']
        cell.fill = header_style['fill']
        cell.alignment = header_style['align']
        cell.border = header_style['border']
    ws.row_dimensions[2].height = 30

    # Sample demo row 3
    demo_row = [
        1, 'SV001', 'Nguyen Van A', '0912345678', 'Nam', '20/10/2002', 'Kinh',
        'Ha Noi', 'Lop truong', 'K63 CNTT', 'Chi bo 1', '01-BC/CB',
        '15/01/2024', 'Tran Van B', '20/01/2024', '123-QD/DU',
        '01/02/2024', '01/02/2024 - 10/02/2024', '15/02/2024', '456/CN',
        'Truong DH', 'Khoa CNTT', 'Chi bo 1', 'Dang uy Truong', 'Tinh uy',
        'DU01', 'Chua ket nap', '', '', '', '', '', '', 'Dang theo doi'
    ]
    ws.append(demo_row)

    # Column widths
    for i, w in enumerate(COL_WIDTHS[:len(HEADERS)], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='File_Mau_Nhap_Du_Lieu_DangVien.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/export/all', methods=['POST'])
def export_all():
    """Loại 1: Xuất danh sách Excel toàn bộ (Excel format)"""
    data = request.json or {}
    ft = data.get('filter_type', 'all')
    fv = data.get('filter_value', '')
    db = get_db()
    try:
        with db.cursor() as cur:
            if ft == 'lop' and fv:
                cur.execute("SELECT * FROM doi_tuong WHERE lop=%s ORDER BY ho_ten", (fv,))
                subtitle = f'Lớp: {fv}'
                fname = f'DanhSach_Lop_{fv.replace(" ","_")}'
            elif ft == 'chibo' and fv:
                cur.execute("SELECT * FROM doi_tuong WHERE chi_bo_cong_nhan=%s ORDER BY ho_ten", (fv,))
                subtitle = f'Chi bộ: {fv}'
                fname = f'DanhSach_ChiBo'
            else:
                cur.execute("SELECT * FROM doi_tuong ORDER BY ho_ten")
                subtitle = 'Toàn trường'
                fname = 'DanhSach_ToanBo'
            rows = cur.fetchall()

        wb = build_list_workbook(rows, subtitle)
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        dl = f'{fname}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return send_file(buf, as_attachment=True, download_name=dl,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    finally:
        db.close()


@app.route('/api/export/single/<int:pid>')
def export_single(pid):
    """Loại 2: Xuất hồ sơ chi tiết cá nhân dưới dạng PDF"""
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT * FROM doi_tuong WHERE id=%s", (pid,))
            row = cur.fetchone()
        if not row:
            return jsonify({'error': 'Not found'}), 404

        pdf_buf = build_profile_pdf(row)
        name = (row.get('ho_ten') or 'HoSo').replace(' ', '_')
        dl = f'HoSo_{name}_{datetime.now().strftime("%Y%m%d")}.pdf'
        return send_file(pdf_buf, as_attachment=True, download_name=dl,
                         mimetype='application/pdf')
    finally:
        db.close()


@app.route('/api/export/selected', methods=['POST'])
def export_selected():
    """Loại 3: Xuất danh sách nhiều người chọn lọc dưới dạng PDF"""
    data = request.json or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'error': 'No IDs'}), 400
    db = get_db()
    try:
        with db.cursor() as cur:
            ph = ','.join(['%s'] * len(ids))
            cur.execute(f"SELECT * FROM doi_tuong WHERE id IN ({ph}) ORDER BY ho_ten", ids)
            rows = cur.fetchall()
        
        pdf_buf = build_list_pdf(rows, f'Danh sách chọn lọc ({len(rows)} người)')
        dl = f'DanhSach_ChonLoc_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
        return send_file(pdf_buf, as_attachment=True, download_name=dl,
                         mimetype='application/pdf')
    finally:
        db.close()


# ─── PDF Form Builders (Mẫu 1-KNĐ, 2-KNĐ, 3-KNĐ, 4-KNĐ, 4a-KNĐ, 5-KNĐ, Mẫu I, Mẫu II) ─
FORM_DOCX_MAP = {
    '1-knd': 'mau-1-KND-don-xin-vao-dang-.docx',
    '2-knd': 'mau-2-KND-ly-lich-vao-dang.docx',
    '3-knd': 'mau-3-KND-giay-gioi-thieu-dang-vien-chinh-thuc.docx',
    '4-knd': 'mau-4-KND-giay-gioi-thieu-doan-vien-uu-tu.docx',
    '4a-knd': 'mau-4a-KND-giay-gioi-thieu-doan-vien-cong-doan.docx',
    '5-knd': 'mau-5-KND-tong-hop-y-kien.docx',
    'mau-i': 'mau-CN-NTVD1.docx',
    'mau-ii': 'mau-CN-NTVD1-2.docx',
}

def build_knd_form_pdf(form_type: str, row: dict) -> io.BytesIO:
    """Render trực tiếp PDF 100% chuẩn thể thức hành chính 2026 bằng ReportLab Engine, dữ liệu bôi viền đỏ nổi bật để người dùng dễ copy/dán"""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=54, rightMargin=54, topMargin=54, bottomMargin=54
    )
    story = []

    def tag_v(val):
        """Hàm bọc dữ liệu động với viền đỏ, chữ đỏ đậm nổi bật"""
        if not val or val == '…………………………………………' or val == '……/……/……':
            val_str = str(val or '…………………………………………')
            return f'<font color="#C8102E"><b>{val_str}</b></font>'
        return f'<font color="#C8102E"><b>[ {val} ]</b></font>'

    ho_ten = tag_v(row.get('ho_ten', '…………………………………………'))
    ngay_sinh = tag_v(fmt_date(row.get('ngay_sinh')) or '……/……/……')
    que_quan = tag_v(row.get('que_quan', '……………………………………………………………………………………'))
    dan_toc = tag_v(row.get('dan_toc', '……………………'))
    lop = tag_v(row.get('lop', '…………………………………………'))
    chuc_vu = tag_v(row.get('chuc_vu', 'Quần chúng'))
    chi_bo = tag_v(row.get('chi_bo_cong_nhan', '…………………………………………'))
    dang_vien_giup_do = tag_v(row.get('dang_vien_giup_do', '…………………………………………'))
    so_qd_cc = tag_v(row.get('so_qd_cc', 'GCN-2026/ĐU'))

    st_header_party = ParagraphStyle('HP', fontName=FONT_BOLD, fontSize=11, leading=14, alignment=0, textColor=colors.HexColor('#C8102E'))
    st_nation_motto = ParagraphStyle('NM', fontName=FONT_BOLD, fontSize=10, leading=13, alignment=1)
    st_title        = ParagraphStyle('T', fontName=FONT_BOLD, fontSize=15, leading=19, alignment=1, textColor=colors.HexColor('#C8102E'))
    st_subtitle     = ParagraphStyle('ST', fontName=FONT_BOLD, fontSize=12, leading=15, alignment=1)
    st_body         = ParagraphStyle('B', fontName=FONT_REGULAR, fontSize=11, leading=18, alignment=4)
    st_body_bold    = ParagraphStyle('BB', fontName=FONT_BOLD, fontSize=11, leading=18, alignment=4)
    st_sign_title   = ParagraphStyle('STT', fontName=FONT_BOLD, fontSize=11, leading=14, alignment=1)
    st_sign_sub     = ParagraphStyle('STS', fontName=FONT_ITALIC, fontSize=10, leading=13, alignment=1)

    if form_type == '1-knd':
        head_tbl = Table([
            [Paragraph("<b>ĐẢNG CỘNG SẢN VIỆT NAM</b>", st_header_party),
             Paragraph("<b>CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM</b><br/><i>Độc lập - Tự do - Hạnh phúc</i>", st_nation_motto)]
        ], colWidths=[240, 264])
        story.append(head_tbl)
        story.append(Spacer(1, 20))

        story.append(Paragraph("DANH SÁCH / ĐƠN XIN VÀO ĐẢNG", st_title))
        story.append(Spacer(1, 15))

        story.append(Paragraph("<b><u>Kính gửi:</u></b> - Chi ủy Chi bộ: " + chi_bo + "<br/>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;- Đảng ủy Trường Đại học Tây Bắc", st_body_bold))
        story.append(Spacer(1, 12))

        p1 = f"""
        Tôi là: {ho_ten}, sinh ngày {ngay_sinh}<br/>
        Nơi sinh: {que_quan}<br/>
        Quê quán: {que_quan}<br/>
        Dân tộc: {dan_toc} &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; Tôn giáo: Không<br/>
        Trình độ học vấn: Đang học Đại học (Lớp {lop})<br/>
        Nơi cư trú: Trường Đại học Tây Bắc<br/>
        Nghề nghiệp: Sinh viên<br/>
        Đơn vị công tác/Học tập: Lớp {lop}, Trường Đại học Tây Bắc<br/>
        Chức vụ chính quyền, đoàn thể: {chuc_vu}
        """
        story.append(Paragraph(p1, st_body))
        story.append(Spacer(1, 12))

        p2 = """
        Sau một thời gian tìm hiểu về Đảng, nghiên cứu Điều lệ Đảng Cộng sản Việt Nam; được sự tuyên truyền, giáo dục và giúp đỡ của Chi bộ, Đoàn thanh niên Trường Đại học Tây Bắc, tôi đã nhận thức sâu sắc về mục đích, lý tưởng cách mạng của Đảng. Tôi tự nguyện xin vào Đảng Cộng sản Việt Nam với mong muốn được cống hiến và trưởng thành.
        """
        story.append(Paragraph(p2, st_body))
        story.append(Spacer(1, 25))

        sign_table = Table([
            ["", Paragraph(f"<i>Tây Bắc, ngày {datetime.now().day} tháng {datetime.now().month} năm {datetime.now().year}</i>", st_sign_sub)],
            ["", Paragraph("<b>NGƯỜI LÀM ĐƠN</b>", st_sign_title)],
            ["", Paragraph("<i>(Ký và ghi rõ họ tên)</i>", st_sign_sub)],
            ["", Spacer(1, 40)],
            ["", Paragraph(f"<b>{ho_ten}</b>", st_sign_title)]
        ], colWidths=[270, 234])
        story.append(sign_table)

    elif form_type == '2-knd':
        story.append(Paragraph("MẪU 2-KNĐ", st_header_party))
        story.append(Spacer(1, 10))
        story.append(Paragraph("LÝ LỊCH CỦA NGƯỜI XIN VÀO ĐẢNG", st_title))
        story.append(Spacer(1, 15))

        p_info = f"""
        Họ và tên khai sinh: {ho_ten}<br/>
        Họ và tên thường dùng: {ho_ten}<br/>
        Ngày, tháng, năm sinh: {ngay_sinh}<br/>
        Quê quán: {que_quan}<br/>
        Dân tộc: {dan_toc} &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; Lớp: {lop}<br/>
        Chi bộ công nhận: {chi_bo}<br/>
        Đảng viên giúp đỡ: {dang_vien_giup_do}
        """
        story.append(Paragraph(p_info, st_body))
        story.append(Spacer(1, 20))
        story.append(Paragraph("<i>(Hồ sơ lý lịch chi tiết được lưu trữ theo quy định của Ban Tổ chức Trung ương năm 2026)</i>", st_sign_sub))

    elif form_type == '3-knd':
        story.append(Paragraph("ĐẢNG CỘNG SẢN VIỆT NAM", st_header_party))
        story.append(Spacer(1, 10))
        story.append(Paragraph("GIẤY GIỚI THIỆU NGƯỜI VÀO ĐẢNG", st_title))
        story.append(Paragraph("(Của Đảng viên chính thức)", st_subtitle))
        story.append(Spacer(1, 15))

        p_intro = f"""
        Tôi là Đảng viên chính thức: {dang_vien_giup_do}<br/>
        Sinh hoạt tại Chi bộ: {chi_bo}<br/><br/>
        Phân công giúp đỡ quần chúng: {ho_ten}, sinh ngày {ngay_sinh}, là học viên/sinh viên Lớp {lop} xin vào Đảng.<br/><br/>
        Qua quá trình theo dõi, giúp đỡ, tôi nhận thấy quần chúng {ho_ten} có lập trường tư tưởng vững vàng, chấp hành tốt chính sách pháp luật, có tinh thần trách nhiệm cao trong học tập và công tác.
        """
        story.append(Paragraph(p_intro, st_body))
        story.append(Spacer(1, 25))

        sign_table = Table([
            ["", Paragraph("<b>ĐẢNG VIÊN GIỚI THIỆU</b>", st_sign_title)],
            ["", Paragraph("<i>(Ký và ghi rõ họ tên)</i>", st_sign_sub)],
            ["", Spacer(1, 40)],
            ["", Paragraph(f"<b>{dang_vien_giup_do}</b>", st_sign_title)]
        ], colWidths=[270, 234])
        story.append(sign_table)

    elif form_type in ['4-knd', '4a-knd']:
        story.append(Paragraph("ĐOÀN TNCS HỒ CHÍ MINH / CÔNG ĐOÀN", st_header_party))
        story.append(Spacer(1, 10))
        story.append(Paragraph("NGHỊ QUYẾT GIỚI THIỆU ĐOÀN VIÊN ƯU TÚ VÀO ĐẢNG", st_title))
        story.append(Spacer(1, 15))

        p_doan = f"""
        Ban Chấp hành Đoàn/Công đoàn Trường Đại học Tây Bắc giới thiệu đoàn viên ưu tú:<br/>
        Họ và tên: {ho_ten} &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; Sinh ngày: {ngay_sinh}<br/>
        Sinh hoạt tại Chi đoàn / Lớp: {lop}<br/>
        Chức vụ: {chuc_vu}<br/><br/>
        Đoàn viên {ho_ten} đạt tiêu chuẩn Đoàn viên ưu tú, tích cực tham gia các phong trào thi đua và có nguyện vọng tha thiết được kết nạp vào Đảng Cộng sản Việt Nam.
        """
        story.append(Paragraph(p_doan, st_body))
        story.append(Spacer(1, 25))

        sign_table = Table([
            ["", Paragraph("<b>T/M BAN CHẤP HÀNH</b>", st_sign_title)],
            ["", Paragraph("BÍ THƯ", st_sign_sub)],
            ["", Spacer(1, 40)],
            ["", Paragraph("<i>(Ký, đóng dấu)</i>", st_sign_sub)]
        ], colWidths=[270, 234])
        story.append(sign_table)

    elif form_type == '5-knd':
        story.append(Paragraph("ĐẢNG CỘNG SẢN VIỆT NAM", st_header_party))
        story.append(Spacer(1, 10))
        story.append(Paragraph("TỔNG HỢP Ý KIẾN NHẬN XÉT CỦA ĐOÀN THỂ VÀ CHI UỶ NƠI CƯ TRÚ", st_title))
        story.append(Spacer(1, 15))

        p_yk = f"""
        Tổng hợp ý kiến nhận xét đối với quần chúng xin vào Đảng: {ho_ten}<br/>
        Sinh ngày: {ngay_sinh} &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; Lớp: {lop}<br/>
        Chi bộ công nhận: {chi_bo}<br/><br/>
        1. Ý kiến của đại diện các đoàn thể chính trị - xã hội nơi làm việc/học tập: Thống nhất cao đề nghị kết nạp.<br/>
        2. Ý kiến của Chi ủy nơi cư trú: Quần chúng và gia đình chấp hành tốt quy định tại địa phương.
        """
        story.append(Paragraph(p_yk, st_body))
        story.append(Spacer(1, 25))

        sign_table = Table([
            ["", Paragraph("<b>T/M CHI UỶ</b>", st_sign_title)],
            ["", Paragraph("BÍ THƯ", st_sign_sub)],
            ["", Spacer(1, 40)],
            ["", Paragraph("<i>(Ký và ghi rõ họ tên)</i>", st_sign_sub)]
        ], colWidths=[270, 234])
        story.append(sign_table)

    elif form_type in ['mau-i', 'mau-ii']:
        story.append(Paragraph("ĐẢNG CỘNG SẢN VIỆT NAM", st_header_party))
        story.append(Spacer(1, 10))
        story.append(Paragraph("GIẤY CHỨNG NHẬN HỌC LỚP BỒI DƯỠNG NHẬN THỨC VỀ ĐẢNG", st_title))
        story.append(Spacer(1, 15))

        p_text = f"""
        Chứng nhận quần chúng: {ho_ten}<br/>
        Sinh ngày: {ngay_sinh} &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; Dân tộc: {dan_toc}<br/>
        Đơn vị công tác/Học tập: Lớp {lop}, Trường Đại học Tây Bắc<br/><br/>
        Đã hoàn thành chương trình <b>Bồi dưỡng nhận thức về Đảng theo quy định mới năm 2026</b>.<br/>
        Kết quả xếp loại: <b>Khá / Giỏi</b><br/>
        Số chứng nhận: {so_qd_cc}
        """
        story.append(Paragraph(p_text, st_body))
        story.append(Spacer(1, 25))

        sign_table = Table([
            ["", Paragraph("<b>T/M BAN THƯỜNG VỤ ĐẢNG UỶ</b>", st_sign_title)],
            ["", Paragraph("BÍ THƯ", st_sign_sub)],
            ["", Spacer(1, 40)],
            ["", Paragraph("<i>(Ký và đóng dấu)</i>", st_sign_sub)]
        ], colWidths=[270, 234])
        story.append(sign_table)

    doc.build(story)
    buf.seek(0)
    return buf


REQUIRED_FORM_FIELDS = {
    '1-knd': [('ho_ten', 'Họ và tên'), ('ngay_sinh', 'Ngày sinh'), ('que_quan', 'Quê quán'), ('dan_toc', 'Dân tộc'), ('lop', 'Lớp')],
    '2-knd': [('ho_ten', 'Họ và tên'), ('ngay_sinh', 'Ngày sinh'), ('que_quan', 'Quê quán'), ('dan_toc', 'Dân tộc'), ('lop', 'Lớp'), ('chi_bo_cong_nhan', 'Chi bộ công nhận'), ('dang_vien_giup_do', 'Đảng viên giúp đỡ')],
    '3-knd': [('ho_ten', 'Họ và tên'), ('chi_bo_cong_nhan', 'Chi bộ công nhận'), ('dang_vien_giup_do', 'Đảng viên giúp đỡ')],
    '4-knd': [('ho_ten', 'Họ và tên'), ('ngay_sinh', 'Ngày sinh'), ('dan_toc', 'Dân tộc'), ('lop', 'Lớp')],
    '4a-knd': [('ho_ten', 'Họ và tên'), ('ngay_sinh', 'Ngày sinh'), ('chuc_vu', 'Chức vụ'), ('chi_bo_cong_nhan', 'Đơn vị Chi bộ')],
    '5-knd': [('ho_ten', 'Họ và tên'), ('chi_bo_cong_nhan', 'Chi bộ công nhận')],
    'mau-i': [('ho_ten', 'Họ và tên'), ('ngay_sinh', 'Ngày sinh'), ('dan_toc', 'Dân tộc'), ('lop', 'Lớp'), ('so_qd_cc', 'Số chứng chỉ')],
    'mau-ii': [('ho_ten', 'Họ và tên'), ('ngay_sinh', 'Ngày sinh'), ('dan_toc', 'Dân tộc'), ('lop', 'Lớp'), ('so_qd_cc', 'Số chứng chỉ')],
}


@app.route('/api/export/form/<form_type>/<int:pid>')
def export_knd_form(form_type, pid):
    """Xuất các Mẫu phiếu chuẩn PDF: 1-knd, 2-knd, 3-knd, 4-knd, 4a-knd, 5-knd, mau-i, mau-ii"""
    db = get_db()
    try:
        with db.cursor() as cur:
            cur.execute("SELECT * FROM doi_tuong WHERE id=%s", (pid,))
            row = cur.fetchone()
        if not row:
            return jsonify({'error': 'Not found'}), 404

        # Check required fields
        req_fields = REQUIRED_FORM_FIELDS.get(form_type, [])
        missing = [label for key, label in req_fields if not row.get(key)]
        if missing:
            return jsonify({
                'error': 'missing_fields',
                'form_type': form_type,
                'doi_tuong_id': pid,
                'ho_ten': row.get('ho_ten', ''),
                'missing_fields': missing
            }), 400

        pdf_buf = build_knd_form_pdf(form_type, row)
        fname = f"Phieu_{form_type.upper()}_{row.get('ho_ten','').replace(' ','_')}.pdf"
        return send_file(pdf_buf, as_attachment=True, download_name=fname, mimetype='application/pdf')
    except Exception as e:
        import traceback
        print("=== ERROR IN export_knd_form ===")
        traceback.print_exc()
        return jsonify({'error': 'internal_server_error', 'details': str(e)}), 500
    finally:
        db.close()


if __name__ == '__main__':
    print("=" * 50)
    print("  Flask API (Excel + PDF) – Hệ thống Kết nạp Đảng v1.1")
    print("  Running at: http://localhost:5000")
    print("  Health check: http://localhost:5000/health")
    print("=" * 50)
    app.run(host='0.0.0.0', port=5000, debug=False)

